from django.test import TestCase
from django.test import override_settings
from django.utils import timezone
from io import BytesIO
from unittest.mock import patch

import openpyxl

from data.models import (
    AssemblyPullData,
    Color,
    Inventory,
    InjectionInventory,
    PositionType,
    Product,
    SafetyStock,
    SystemParameter,
    VehicleModel,
)
from schedule.models import FormationSlot, ScheduleRecord
from schedule.services.algorithms import SchedulingAlgorithm
from schedule.utils import export_schedule_to_excel


class SchedulingRuleAlignmentTests(TestCase):
    def setUp(self):
        self.front = PositionType.objects.create(name="front")
        self.rear = PositionType.objects.create(name="rear")
        self.red = Color.objects.create(name="red")
        self.blue = Color.objects.create(name="blue")
        self.white = Color.objects.create(name="white")
        self.a0 = VehicleModel.objects.create(name="A0")
        self.a1 = VehicleModel.objects.create(name="A1")

        defaults = {
            "CYCLE_TIME_MIN": "300",
            "AVG_HANGING_COUNT": "1",
            "TOTAL_VEHICLES": "10",
            "SHORT_TERM_CAPACITY": "40",
            "LONG_TERM_CAPACITY": "60",
            "FRONT_REAR_BALANCE_D": "1",
            "GROUP_CAPACITY_LIMIT": "50",
            "LONG_TERM_FORECAST_HOURS": "0",
        }
        for key, value in defaults.items():
            SystemParameter.objects.update_or_create(
                param_key=key,
                defaults={"param_value": value},
            )

    def create_product(
        self,
        model,
        color,
        position,
        hanging_count=1,
        yield_rate=100,
        inventory=0,
        injection_inventory=0,
        safety_stock=0,
    ):
        product = Product.objects.create(
            vehicle_model=model,
            color=color,
            position_type=position,
            hanging_count_per_vehicle=hanging_count,
            yield_rate=yield_rate,
        )
        Inventory.objects.create(
            product=product,
            current_quantity=inventory,
        )
        InjectionInventory.objects.create(
            product=product,
            current_quantity=injection_inventory,
        )
        SafetyStock.objects.create(product=product, quantity=safety_stock)
        return product

    def add_pull(self, sequence, model, color):
        AssemblyPullData.objects.create(
            sequence=sequence,
            vehicle_model=model,
            color=color,
            planned_time=timezone.now(),
        )

    def create_record(self):
        return ScheduleRecord.objects.create(
            short_term_duration=2,
            long_term_duration=2,
            total_vehicles=10,
            status="pending",
            cycle_time_min=300,
            avg_hanging_count=1,
            total_vehicles_in_line=10,
            short_term_capacity=40,
            long_term_capacity=60,
            front_rear_balance_d=1,
            group_capacity_limit=50,
        )

    def test_manual_durations_define_short_and_long_read_windows(self):
        red_front = self.create_product(self.a0, self.red, self.front)
        red_rear = self.create_product(self.a0, self.red, self.rear)
        blue_front = self.create_product(self.a0, self.blue, self.front)
        blue_rear = self.create_product(self.a0, self.blue, self.rear)
        white_front = self.create_product(self.a1, self.white, self.front)
        white_rear = self.create_product(self.a1, self.white, self.rear)

        self.add_pull(1, self.a0, self.red)
        self.add_pull(2, self.a0, self.blue)
        self.add_pull(3, self.a0, self.red)
        self.add_pull(4, self.a0, self.blue)
        self.add_pull(5, self.a1, self.white)

        algorithm = SchedulingAlgorithm(short_term_duration=2, long_term_duration=2)
        results = algorithm.calculate()

        short_quantities = {
            item["product"].id: item["demand_quantity"] for item in results["short_term"]
        }
        long_quantities = {
            item["product"].id: item["demand_quantity"] for item in results["long_term"]
        }

        self.assertEqual(short_quantities[red_front.id], 1)
        self.assertEqual(short_quantities[red_rear.id], 1)
        self.assertEqual(short_quantities[blue_front.id], 1)
        self.assertEqual(short_quantities[blue_rear.id], 1)
        self.assertNotIn(white_front.id, short_quantities)
        self.assertNotIn(white_rear.id, short_quantities)

        self.assertEqual(long_quantities[red_front.id], 1)
        self.assertEqual(long_quantities[red_rear.id], 1)
        self.assertEqual(long_quantities[blue_front.id], 1)
        self.assertEqual(long_quantities[blue_rear.id], 1)
        self.assertNotIn(white_front.id, long_quantities)
        self.assertNotIn(white_rear.id, long_quantities)

    def test_current_inventory_is_used_as_effective_inventory(self):
        # The algorithm reads current_quantity directly from Inventory.
        # Inventory of 5 - production_quantity of 1 = final_value of 4.
        red_front = self.create_product(
            self.a0,
            self.red,
            self.front,
            inventory=5,
        )
        self.create_product(self.a0, self.red, self.rear)
        self.add_pull(1, self.a0, self.red)

        algorithm = SchedulingAlgorithm(short_term_duration=1, long_term_duration=0)
        short_risk = algorithm.calculate()["short_risk"]
        risk_by_product = {item["product"].id: item for item in short_risk}

        self.assertEqual(risk_by_product[red_front.id]["final_value"], 4)

    def test_long_term_plan_respects_balance_and_model_capacity_limit(self):
        red_front = self.create_product(
            self.a0,
            self.red,
            self.front,
            safety_stock=10,
            injection_inventory=20,
        )
        red_rear = self.create_product(
            self.a0,
            self.red,
            self.rear,
            safety_stock=10,
            injection_inventory=20,
        )
        blue_front = self.create_product(
            self.a0,
            self.blue,
            self.front,
            safety_stock=10,
            injection_inventory=20,
        )
        blue_rear = self.create_product(
            self.a0,
            self.blue,
            self.rear,
            safety_stock=10,
            injection_inventory=20,
        )

        algorithm = SchedulingAlgorithm(short_term_duration=0, long_term_duration=0)
        long_risks = [
            {
                "product": red_rear,
                "final_value": -3,
                "safety_stock": 10,
                "risk_value": 13,
                "group_risk_value": 13,
            },
            {
                "product": red_front,
                "final_value": -2,
                "safety_stock": 10,
                "risk_value": 12,
                "group_risk_value": 13,
            },
            {
                "product": blue_rear,
                "final_value": -4,
                "safety_stock": 10,
                "risk_value": 14,
                "group_risk_value": 14,
            },
            {
                "product": blue_front,
                "final_value": -4,
                "safety_stock": 10,
                "risk_value": 14,
                "group_risk_value": 14,
            },
        ]

        plan = algorithm.calculate_long_term_plan(long_risks)
        totals_by_position = {"front": 0, "rear": 0}
        total_a0 = 0

        for item in plan:
            totals_by_position[item["product"].position_type.name] += item["vehicle_count"]
            if item["product"].vehicle_model_id == self.a0.id:
                total_a0 += item["vehicle_count"]

        self.assertLessEqual(abs(totals_by_position["front"] - totals_by_position["rear"]), 1)
        self.assertLessEqual(total_a0, 3)

    def test_optimize_formation_reuses_matching_previous_slots(self):
        reusable = self.create_product(self.a0, self.red, self.front)
        other = self.create_product(self.a1, self.blue, self.rear)

        previous = self.create_record()
        previous.status = "completed"
        previous.save(update_fields=["status"])
        FormationSlot.objects.create(
            record=previous,
            slot_number=1,
            product=reusable,
            plan_type="short",
        )
        FormationSlot.objects.create(
            record=previous,
            slot_number=2,
            product=other,
            plan_type="long",
        )

        algorithm = SchedulingAlgorithm(short_term_duration=0, long_term_duration=0)
        formation = algorithm.optimize_formation(
            short_plan=[{"product": reusable, "vehicle_count": 1, "note": "short"}],
            long_plan=[{"product": other, "vehicle_count": 1, "note": "long"}],
        )

        reused_slot = next(slot for slot in formation if slot["product"].id == reusable.id)
        self.assertEqual(reused_slot["slot_number"], 1)
        self.assertTrue(reused_slot["is_reused"])

    def test_calculate_backfills_formation_to_total_vehicles_when_inventory_allows(self):
        self.create_product(
            self.a0,
            self.red,
            self.front,
            hanging_count=1,
            yield_rate=100,
            inventory=0,
            injection_inventory=20,
            safety_stock=5,
        )
        self.create_product(
            self.a0,
            self.red,
            self.rear,
            hanging_count=1,
            yield_rate=100,
            inventory=0,
            injection_inventory=20,
            safety_stock=5,
        )
        self.add_pull(1, self.a0, self.red)

        algorithm = SchedulingAlgorithm(short_term_duration=1, long_term_duration=0)
        results = algorithm.calculate()

        self.assertEqual(len(results["formation"]), 10)
        self.assertTrue(all(slot["product"] is not None for slot in results["formation"]))

    def test_inventory_updates_become_current_inventory_for_next_run(self):
        front_product = self.create_product(
            self.a0,
            self.red,
            self.front,
            hanging_count=2,
            yield_rate=100,
            inventory=5,
            injection_inventory=10,
        )
        self.create_product(
            self.a0,
            self.red,
            self.rear,
            hanging_count=2,
            yield_rate=100,
            inventory=5,
            injection_inventory=10,
        )
        self.add_pull(1, self.a0, self.red)
        self.add_pull(2, self.a0, self.red)

        algorithm = SchedulingAlgorithm(short_term_duration=1, long_term_duration=1)
        results = algorithm.calculate()
        record = self.create_record()

        algorithm.save_results(results, record)

        # After save_results, the Inventory.current_quantity should be updated directly.
        inventory = Inventory.objects.get(product=front_product)
        self.assertIsNotNone(inventory.current_quantity)

        # A subsequent run should read the updated current_quantity.
        next_run = SchedulingAlgorithm(short_term_duration=0, long_term_duration=0)
        key = next_run._get_product_key(front_product)
        self.assertEqual(
            next_run.inventory_data[key]["current"],
            inventory.current_quantity,
        )

    def test_short_term_plan_is_capped_by_injection_inventory(self):
        front_product = self.create_product(
            self.a0,
            self.red,
            self.front,
            hanging_count=2,
            yield_rate=100,
            inventory=0,
            injection_inventory=3,
            safety_stock=0,
        )
        self.create_product(
            self.a0,
            self.red,
            self.rear,
            hanging_count=2,
            yield_rate=100,
            inventory=10,
            injection_inventory=10,
            safety_stock=0,
        )
        self.add_pull(1, self.a0, self.red)
        self.add_pull(2, self.a0, self.red)
        self.add_pull(3, self.a0, self.red)

        algorithm = SchedulingAlgorithm(short_term_duration=3, long_term_duration=0)
        results = algorithm.calculate()
        short_plan = {item["product"].id: item for item in results["short_plan"]}

        self.assertEqual(short_plan[front_product.id]["vehicle_count"], 1)
        self.assertIn("受注塑库存限制", short_plan[front_product.id]["note"])

    def test_long_term_plan_consumes_only_remaining_injection_inventory_after_short_term(self):
        front_product = self.create_product(
            self.a0,
            self.red,
            self.front,
            hanging_count=2,
            yield_rate=100,
            inventory=0,
            injection_inventory=2,
            safety_stock=5,
        )
        self.create_product(
            self.a0,
            self.red,
            self.rear,
            hanging_count=2,
            yield_rate=100,
            inventory=10,
            injection_inventory=10,
            safety_stock=0,
        )
        self.add_pull(1, self.a0, self.red)
        self.add_pull(2, self.a0, self.red)
        self.add_pull(3, self.a0, self.red)

        algorithm = SchedulingAlgorithm(short_term_duration=2, long_term_duration=1)
        results = algorithm.calculate()
        short_plan = {item["product"].id: item for item in results["short_plan"]}
        long_plan = {item["product"].id: item for item in results["long_plan"]}

        self.assertEqual(short_plan[front_product.id]["vehicle_count"], 1)
        self.assertEqual(long_plan[front_product.id]["vehicle_count"], 0)
        self.assertIn("受注塑库存限制", long_plan[front_product.id]["note"])

    def test_excel_export_includes_plan_notes_slot_reuse_and_inventory_snapshots(self):
        reusable = self.create_product(
            self.a0,
            self.red,
            self.front,
            hanging_count=2,
            yield_rate=100,
            inventory=3,
            injection_inventory=8,
        )
        self.create_product(
            self.a0,
            self.red,
            self.rear,
            hanging_count=2,
            yield_rate=100,
            inventory=3,
            injection_inventory=8,
        )
        previous = self.create_record()
        previous.status = "completed"
        previous.save(update_fields=["status"])
        FormationSlot.objects.create(
            record=previous,
            slot_number=1,
            product=reusable,
            plan_type="short",
        )

        self.add_pull(1, self.a0, self.red)
        algorithm = SchedulingAlgorithm(short_term_duration=1, long_term_duration=0)
        record = self.create_record()
        results = algorithm.calculate()
        algorithm.save_results(results, record)

        response = export_schedule_to_excel(record.id)
        workbook = openpyxl.load_workbook(BytesIO(response.content))

        self.assertIn("短期计划", workbook.sheetnames)
        self.assertIn("阵型排布", workbook.sheetnames)
        self.assertIn("涂装库存更新", workbook.sheetnames)
        self.assertIn("注塑库存更新", workbook.sheetnames)

        short_plan_headers = [cell.value for cell in workbook["短期计划"][1]]
        formation_headers = [cell.value for cell in workbook["阵型排布"][1]]
        paint_headers = [cell.value for cell in workbook["涂装库存更新"][1]]

        self.assertIn("计划说明", short_plan_headers)
        self.assertIn("是否复用上一轮", formation_headers)
        self.assertEqual(paint_headers, ["车型", "颜色", "位置", "计算前", "变动", "更新后"])

    def test_history_page_shows_windows_reused_slots_and_inventory_summary(self):
        reusable = self.create_product(
            self.a0,
            self.red,
            self.front,
            hanging_count=2,
            yield_rate=100,
            inventory=3,
            injection_inventory=8,
        )
        self.create_product(
            self.a0,
            self.red,
            self.rear,
            hanging_count=2,
            yield_rate=100,
            inventory=3,
            injection_inventory=8,
        )
        previous = self.create_record()
        previous.status = "completed"
        previous.save(update_fields=["status"])
        FormationSlot.objects.create(
            record=previous,
            slot_number=1,
            product=reusable,
            plan_type="short",
        )

        algorithm = SchedulingAlgorithm(short_term_duration=1, long_term_duration=1)
        results = algorithm.calculate()
        record = self.create_record()
        algorithm.save_results(results, record)
        record.status = "completed"
        record.save(update_fields=["status"])

        response = self.client.get("/schedule/history/")

        self.assertContains(response, "1-2")
        self.assertContains(response, "复用")
        self.assertContains(response, "涂装更新")

    def test_history_page_shows_injection_constraint_summary(self):
        self.create_product(
            self.a0,
            self.red,
            self.front,
            hanging_count=2,
            yield_rate=100,
            inventory=0,
            injection_inventory=3,
            safety_stock=0,
        )
        self.create_product(
            self.a0,
            self.red,
            self.rear,
            hanging_count=2,
            yield_rate=100,
            inventory=10,
            injection_inventory=10,
            safety_stock=0,
        )
        self.add_pull(1, self.a0, self.red)
        self.add_pull(2, self.a0, self.red)
        self.add_pull(3, self.a0, self.red)

        algorithm = SchedulingAlgorithm(short_term_duration=3, long_term_duration=0)
        results = algorithm.calculate()
        record = self.create_record()
        algorithm.save_results(results, record)
        record.status = "completed"
        record.save(update_fields=["status"])

        response = self.client.get("/schedule/history/")

        self.assertContains(response, "注塑受限")
        self.assertContains(response, "1 项")
        self.assertContains(response, "截断 1 车")

    def test_history_page_shows_overview_cards_for_constraints_and_gaps(self):
        SystemParameter.objects.filter(param_key="SHORT_TERM_CAPACITY").update(param_value="10")
        self.create_product(
            self.a0,
            self.red,
            self.front,
            inventory=0,
            injection_inventory=3,
            safety_stock=0,
        )
        self.create_product(
            self.a0,
            self.red,
            self.rear,
            inventory=10,
            injection_inventory=10,
            safety_stock=0,
        )
        self.add_pull(1, self.a0, self.red)
        self.add_pull(2, self.a0, self.red)
        self.add_pull(3, self.a0, self.red)

        algorithm = SchedulingAlgorithm(short_term_duration=3, long_term_duration=0)
        results = algorithm.calculate()
        record = self.create_record()
        algorithm.save_results(results, record)
        record.status = "completed"
        record.save(update_fields=["status"])

        response = self.client.get("/schedule/history/")

        self.assertContains(response, "累计注塑截断")
        self.assertContains(response, "累计计划缺口")
        self.assertContains(response, "最近异常记录")
        self.assertContains(response, f"#{record.id}")

    def test_history_page_shows_demand_and_inventory_item_summary(self):
        self.create_product(
            self.a0,
            self.red,
            self.front,
            inventory=2,
            injection_inventory=5,
            safety_stock=1,
        )
        self.create_product(
            self.a0,
            self.red,
            self.rear,
            inventory=2,
            injection_inventory=5,
            safety_stock=1,
        )
        self.add_pull(1, self.a0, self.red)
        self.add_pull(2, self.a0, self.red)

        algorithm = SchedulingAlgorithm(short_term_duration=1, long_term_duration=1)
        results = algorithm.calculate()
        record = self.create_record()
        algorithm.save_results(results, record)
        record.status = "completed"
        record.save(update_fields=["status"])

        response = self.client.get("/schedule/history/")

        self.assertContains(response, "总成拉动")
        self.assertContains(response, "短期 2 / 长期 2")
        self.assertContains(response, "库存项")
        self.assertContains(response, "涂装 2 / 注塑 2")

    def test_history_page_total_vehicles_uses_actual_formation_slots(self):
        self.create_product(
            self.a0,
            self.red,
            self.front,
            inventory=0,
            injection_inventory=20,
            safety_stock=5,
        )
        self.create_product(
            self.a0,
            self.red,
            self.rear,
            inventory=0,
            injection_inventory=20,
            safety_stock=5,
        )
        self.add_pull(1, self.a0, self.red)

        algorithm = SchedulingAlgorithm(short_term_duration=1, long_term_duration=0)
        results = algorithm.calculate()
        record = self.create_record()
        algorithm.save_results(results, record)
        record.status = "completed"
        record.save(update_fields=["status"])

        response = self.client.get("/schedule/history/")

        self.assertContains(response, "<td>10</td>", html=True)

    def test_history_page_can_delete_record(self):
        record = self.create_record()
        record.status = "completed"
        record.save(update_fields=["status"])

        response = self.client.post(f"/schedule/history/{record.id}/delete/", follow=True)

        self.assertEqual(response.status_code, 200)
        self.assertFalse(ScheduleRecord.objects.filter(id=record.id).exists())
        self.assertContains(response, "历史记录已删除")

    def test_result_page_fills_full_slot_count_when_inventory_allows(self):
        self.create_product(
            self.a0,
            self.red,
            self.front,
            hanging_count=1,
            yield_rate=100,
            inventory=0,
            injection_inventory=20,
            safety_stock=5,
        )
        self.create_product(
            self.a0,
            self.red,
            self.rear,
            hanging_count=1,
            yield_rate=100,
            inventory=0,
            injection_inventory=20,
            safety_stock=5,
        )
        self.create_product(
            self.a1,
            self.blue,
            self.front,
            hanging_count=1,
            yield_rate=100,
            inventory=0,
            injection_inventory=20,
            safety_stock=5,
        )
        self.create_product(
            self.a1,
            self.blue,
            self.rear,
            hanging_count=1,
            yield_rate=100,
            inventory=0,
            injection_inventory=20,
            safety_stock=5,
        )
        self.add_pull(1, self.a0, self.red)
        self.add_pull(2, self.a1, self.blue)

        algorithm = SchedulingAlgorithm(short_term_duration=1, long_term_duration=0)
        results = algorithm.calculate()
        record = self.create_record()
        algorithm.save_results(results, record)
        record.status = "completed"
        record.save(update_fields=["status"])

        response = self.client.get(f"/schedule/result/{record.id}/")

        self.assertContains(response, "槽位号")
        self.assertContains(response, "空槽位")
        self.assertContains(response, "<td>10</td>", html=True)
        self.assertContains(response, "暂无长期组风险摘要")
        self.assertContains(response, "短期")
        self.assertContains(response, "长期")

    def test_result_page_filters_formation_slots_by_vehicle_color_and_position(self):
        self.create_product(
            self.a0,
            self.red,
            self.front,
            hanging_count=1,
            yield_rate=100,
            inventory=0,
            injection_inventory=20,
            safety_stock=5,
        )
        self.create_product(
            self.a0,
            self.red,
            self.rear,
            hanging_count=1,
            yield_rate=100,
            inventory=0,
            injection_inventory=20,
            safety_stock=5,
        )
        self.create_product(
            self.a1,
            self.blue,
            self.front,
            hanging_count=1,
            yield_rate=100,
            inventory=0,
            injection_inventory=20,
            safety_stock=5,
        )
        self.create_product(
            self.a1,
            self.blue,
            self.rear,
            hanging_count=1,
            yield_rate=100,
            inventory=0,
            injection_inventory=20,
            safety_stock=5,
        )
        self.add_pull(1, self.a0, self.red)
        self.add_pull(2, self.a1, self.blue)

        algorithm = SchedulingAlgorithm(short_term_duration=2, long_term_duration=0)
        results = algorithm.calculate()
        record = self.create_record()
        algorithm.save_results(results, record)
        record.status = "completed"
        record.save(update_fields=["status"])

        response = self.client.get(f"/schedule/result/{record.id}/?vehicle=A0&color=red&position=rear")

        self.assertContains(response, "A0")
        self.assertContains(response, "红")
        self.assertContains(response, "后")
        filtered_slots = response.context["formation_slots"]
        self.assertTrue(filtered_slots)
        self.assertTrue(all(
            slot.product.vehicle_model.name == "A0"
            and slot.product.color.name == "red"
            and slot.product.position_type.name == "rear"
            for slot in filtered_slots
            if getattr(slot, "product", None)
        ))

    def test_long_term_plan_note_explains_priority_and_constraints(self):
        red_front = self.create_product(
            self.a0,
            self.red,
            self.front,
            safety_stock=10,
            injection_inventory=20,
        )
        red_rear = self.create_product(
            self.a0,
            self.red,
            self.rear,
            safety_stock=10,
            injection_inventory=20,
        )
        blue_front = self.create_product(
            self.a0,
            self.blue,
            self.front,
            safety_stock=10,
            injection_inventory=20,
        )
        blue_rear = self.create_product(
            self.a0,
            self.blue,
            self.rear,
            safety_stock=10,
            injection_inventory=20,
        )

        algorithm = SchedulingAlgorithm(short_term_duration=0, long_term_duration=0)
        long_risks = [
            {
                "product": red_rear,
                "final_value": -3,
                "safety_stock": 10,
                "risk_value": 13,
                "group_risk_value": 13,
            },
            {
                "product": red_front,
                "final_value": 0,
                "safety_stock": 10,
                "risk_value": 10,
                "group_risk_value": 13,
            },
            {
                "product": blue_front,
                "final_value": -4,
                "safety_stock": 10,
                "risk_value": 14,
                "group_risk_value": 14,
            },
            {
                "product": blue_rear,
                "final_value": -4,
                "safety_stock": 10,
                "risk_value": 14,
                "group_risk_value": 14,
            },
        ]

        plan = algorithm.calculate_long_term_plan(long_risks)
        all_notes = " ".join(item["note"] for item in plan)

        self.assertTrue("优先补前件" in all_notes or "优先补后件" in all_notes)
        self.assertIn("车型上限", all_notes)

    def test_long_term_plan_keeps_capacity_limit_reason_for_unallocated_items(self):
        SystemParameter.objects.filter(param_key="LONG_TERM_CAPACITY").update(param_value="10")
        red_front = self.create_product(
            self.a0,
            self.red,
            self.front,
            safety_stock=10,
            injection_inventory=20,
        )
        red_rear = self.create_product(
            self.a0,
            self.red,
            self.rear,
            safety_stock=10,
            injection_inventory=20,
        )
        blue_front = self.create_product(
            self.a0,
            self.blue,
            self.front,
            safety_stock=10,
            injection_inventory=20,
        )
        blue_rear = self.create_product(
            self.a0,
            self.blue,
            self.rear,
            safety_stock=10,
            injection_inventory=20,
        )

        algorithm = SchedulingAlgorithm(short_term_duration=0, long_term_duration=0)
        long_risks = [
            {
                "product": red_rear,
                "final_value": -3,
                "safety_stock": 10,
                "risk_value": 13,
                "group_risk_value": 13,
            },
            {
                "product": red_front,
                "final_value": -3,
                "safety_stock": 10,
                "risk_value": 13,
                "group_risk_value": 13,
            },
            {
                "product": blue_rear,
                "final_value": -4,
                "safety_stock": 10,
                "risk_value": 14,
                "group_risk_value": 14,
            },
            {
                "product": blue_front,
                "final_value": -4,
                "safety_stock": 10,
                "risk_value": 14,
                "group_risk_value": 14,
            },
        ]

        plan = algorithm.calculate_long_term_plan(long_risks)
        all_notes = " ".join(item["note"] for item in plan)

        self.assertIn("受长期产能限制", all_notes)

    def test_result_page_shows_long_term_capacity_reason_in_plan_gap_summary(self):
        SystemParameter.objects.filter(param_key="LONG_TERM_CAPACITY").update(param_value="10")
        self.create_product(
            self.a0,
            self.red,
            self.front,
            inventory=0,
            injection_inventory=20,
            safety_stock=10,
        )
        self.create_product(
            self.a0,
            self.red,
            self.rear,
            inventory=0,
            injection_inventory=20,
            safety_stock=10,
        )
        self.create_product(
            self.a0,
            self.blue,
            self.front,
            inventory=0,
            injection_inventory=20,
            safety_stock=10,
        )
        self.create_product(
            self.a0,
            self.blue,
            self.rear,
            inventory=0,
            injection_inventory=20,
            safety_stock=10,
        )
        self.add_pull(1, self.a0, self.red)
        self.add_pull(2, self.a0, self.blue)

        algorithm = SchedulingAlgorithm(short_term_duration=0, long_term_duration=2)
        results = algorithm.calculate()
        record = self.create_record()
        algorithm.save_results(results, record)
        record.status = "completed"
        record.save(update_fields=["status"])

        response = self.client.get(f"/schedule/result/{record.id}/")

        self.assertContains(response, "计划缺口摘要")
        self.assertContains(response, "长期")
        self.assertContains(response, "受长期产能限制")

    def test_calculate_page_shows_parameter_derivation_and_risk_preview(self):
        self.create_product(
            self.a0,
            self.red,
            self.front,
            inventory=1,
            safety_stock=3,
        )
        self.create_product(
            self.a0,
            self.red,
            self.rear,
            inventory=5,
            safety_stock=3,
        )
        self.add_pull(1, self.a0, self.red)
        self.add_pull(2, self.a0, self.blue)

        response = self.client.get("/schedule/calculate/")

        self.assertContains(response, "推荐短期覆盖")
        self.assertContains(response, "当前总成拉动")
        self.assertContains(response, "低于安全库存")

    def test_result_page_shows_summary_cards_for_risk_and_formation(self):
        self.create_product(
            self.a0,
            self.red,
            self.front,
            hanging_count=1,
            yield_rate=100,
            inventory=0,
            injection_inventory=2,
            safety_stock=5,
        )
        self.create_product(
            self.a0,
            self.red,
            self.rear,
            hanging_count=1,
            yield_rate=100,
            inventory=0,
            injection_inventory=2,
            safety_stock=5,
        )
        self.add_pull(1, self.a0, self.red)

        algorithm = SchedulingAlgorithm(short_term_duration=1, long_term_duration=0)
        results = algorithm.calculate()
        record = self.create_record()
        algorithm.save_results(results, record)
        record.status = "completed"
        record.save(update_fields=["status"])

        response = self.client.get(f"/schedule/result/{record.id}/")

        self.assertContains(response, "短期缺货物料")
        self.assertContains(response, "长期风险组")
        self.assertContains(response, "复用槽位")
        self.assertContains(response, "空槽位")

    def test_result_page_shows_long_risk_group_summary(self):
        self.create_product(
            self.a0,
            self.red,
            self.front,
            hanging_count=1,
            yield_rate=100,
            inventory=0,
            injection_inventory=2,
            safety_stock=5,
        )
        self.create_product(
            self.a0,
            self.red,
            self.rear,
            hanging_count=1,
            yield_rate=100,
            inventory=0,
            injection_inventory=2,
            safety_stock=5,
        )
        self.add_pull(1, self.a0, self.red)
        self.add_pull(2, self.a0, self.red)
        self.add_pull(3, self.a0, self.red)

        algorithm = SchedulingAlgorithm(short_term_duration=0, long_term_duration=3)
        results = algorithm.calculate()
        record = self.create_record()
        algorithm.save_results(results, record)
        record.status = "completed"
        record.save(update_fields=["status"])

        response = self.client.get(f"/schedule/result/{record.id}/")

        self.assertContains(response, "长期组风险摘要")
        self.assertContains(response, "A0 / 红")
        self.assertContains(response, "高风险组")

    def test_color_model_exposes_chinese_display_name(self):
        self.assertEqual(self.red.display_name, "红")
        self.assertEqual(self.blue.display_name, "蓝")
        self.assertEqual(self.white.display_name, "白")

    def test_result_page_shows_injection_constraint_summary(self):
        self.create_product(
            self.a0,
            self.red,
            self.front,
            hanging_count=2,
            yield_rate=100,
            inventory=0,
            injection_inventory=3,
            safety_stock=5,
        )
        self.create_product(
            self.a0,
            self.red,
            self.rear,
            hanging_count=2,
            yield_rate=100,
            inventory=10,
            injection_inventory=10,
            safety_stock=0,
        )
        self.add_pull(1, self.a0, self.red)
        self.add_pull(2, self.a0, self.red)
        self.add_pull(3, self.a0, self.red)

        algorithm = SchedulingAlgorithm(short_term_duration=3, long_term_duration=0)
        results = algorithm.calculate()
        record = self.create_record()
        algorithm.save_results(results, record)
        record.status = "completed"
        record.save(update_fields=["status"])

        response = self.client.get(f"/schedule/result/{record.id}/")

        self.assertContains(response, "注塑受限物料")
        self.assertContains(response, "注塑约束摘要")
        self.assertContains(response, "A0")
        self.assertContains(response, "短期")
        self.assertContains(response, "截断 1 车")

    def test_excel_export_summary_includes_injection_constraint_metrics(self):
        self.create_product(
            self.a0,
            self.red,
            self.front,
            hanging_count=2,
            yield_rate=100,
            inventory=0,
            injection_inventory=3,
            safety_stock=0,
        )
        self.create_product(
            self.a0,
            self.red,
            self.rear,
            hanging_count=2,
            yield_rate=100,
            inventory=10,
            injection_inventory=10,
            safety_stock=0,
        )
        self.add_pull(1, self.a0, self.red)
        self.add_pull(2, self.a0, self.red)
        self.add_pull(3, self.a0, self.red)

        algorithm = SchedulingAlgorithm(short_term_duration=3, long_term_duration=0)
        results = algorithm.calculate()
        record = self.create_record()
        algorithm.save_results(results, record)

        response = export_schedule_to_excel(record.id)
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        summary_sheet = workbook["计算摘要"]
        summary_values = {
            summary_sheet.cell(row=row, column=1).value: summary_sheet.cell(row=row, column=2).value
            for row in range(2, summary_sheet.max_row + 1)
        }

        self.assertEqual(summary_values["注塑受限物料数"], 1)
        self.assertEqual(summary_values["注塑截断车数"], 1)

    def test_result_page_shows_plan_gap_summary_with_unallocated_reason(self):
        SystemParameter.objects.filter(param_key="SHORT_TERM_CAPACITY").update(param_value="10")
        self.create_product(
            self.a0,
            self.red,
            self.front,
            inventory=0,
            injection_inventory=10,
            safety_stock=0,
        )
        self.create_product(
            self.a0,
            self.red,
            self.rear,
            inventory=0,
            injection_inventory=10,
            safety_stock=0,
        )
        self.create_product(
            self.a0,
            self.blue,
            self.front,
            inventory=0,
            injection_inventory=10,
            safety_stock=0,
        )
        self.create_product(
            self.a0,
            self.blue,
            self.rear,
            inventory=0,
            injection_inventory=10,
            safety_stock=0,
        )
        self.add_pull(1, self.a0, self.red)
        self.add_pull(2, self.a0, self.blue)

        algorithm = SchedulingAlgorithm(short_term_duration=2, long_term_duration=0)
        results = algorithm.calculate()
        record = self.create_record()
        algorithm.save_results(results, record)
        record.status = "completed"
        record.save(update_fields=["status"])

        response = self.client.get(f"/schedule/result/{record.id}/")

        self.assertContains(response, "计划缺口摘要")
        self.assertContains(response, "未满足需求")
        self.assertContains(response, "受短期产能限制")

    def test_excel_export_includes_plan_gap_sheet(self):
        SystemParameter.objects.filter(param_key="SHORT_TERM_CAPACITY").update(param_value="10")
        self.create_product(
            self.a0,
            self.red,
            self.front,
            inventory=0,
            injection_inventory=10,
            safety_stock=0,
        )
        self.create_product(
            self.a0,
            self.red,
            self.rear,
            inventory=0,
            injection_inventory=10,
            safety_stock=0,
        )
        self.create_product(
            self.a0,
            self.blue,
            self.front,
            inventory=0,
            injection_inventory=10,
            safety_stock=0,
        )
        self.create_product(
            self.a0,
            self.blue,
            self.rear,
            inventory=0,
            injection_inventory=10,
            safety_stock=0,
        )
        self.add_pull(1, self.a0, self.red)
        self.add_pull(2, self.a0, self.blue)

        algorithm = SchedulingAlgorithm(short_term_duration=2, long_term_duration=0)
        results = algorithm.calculate()
        record = self.create_record()
        algorithm.save_results(results, record)

        response = export_schedule_to_excel(record.id)
        workbook = openpyxl.load_workbook(BytesIO(response.content))

        self.assertIn("计划缺口", workbook.sheetnames)
        gap_headers = [cell.value for cell in workbook["计划缺口"][1]]
        self.assertEqual(
            gap_headers,
            ["阶段", "车型", "颜色", "位置", "需要车数", "实际分配", "缺口车数", "未满足需求", "原因"],
        )


class MigrationGuardTests(TestCase):
    @override_settings(ALLOWED_HOSTS=["testserver"])
    @patch("painting.middleware.has_pending_migrations", return_value=True)
    def test_request_is_blocked_with_clear_message_when_migrations_are_pending(self, _mock_pending):
        response = self.client.get("/")

        self.assertEqual(response.status_code, 503)
        self.assertContains(response, "检测到未执行的数据库迁移", status_code=503)
        self.assertContains(response, "python manage.py migrate", status_code=503)
